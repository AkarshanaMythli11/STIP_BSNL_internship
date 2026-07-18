import os
import sqlite3
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from werkzeug.security import check_password_hash
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from datetime import datetime

# Initialize Flask App
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'stip_secret_key_bsnl_2026')

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'database.db')
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'uploads')
REPORTS_FOLDER = os.path.join(BASE_DIR, 'reports')

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(REPORTS_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB limit

# Helper: Connect to Database
def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# Helper: Calculate TIPI score and Category
def calculate_tipi(population, signal_strength, distance, internet_users, complaints, fiber, network_type):
    try:
        population = int(population)
        signal_strength = int(signal_strength)
        distance = float(distance)
        internet_users = int(internet_users)
        complaints = int(complaints)
    except (ValueError, TypeError):
        # Fallback default values
        population = 0
        signal_strength = 3
        distance = 0.0
        internet_users = 0
        complaints = 0

    fiber_clean = str(fiber).strip().lower()
    net_clean = str(network_type).strip().upper()

    # Sub-scores
    s_pop = min(population / 50000.0, 1.0) * 100.0
    s_sig = (1.0 - (signal_strength - 1) / 4.0) * 100.0 if 1 <= signal_strength <= 5 else 50.0
    s_dist = min(distance / 10.0, 1.0) * 100.0
    
    net_scores = {"2G": 100.0, "3G": 75.0, "4G": 40.0, "5G": 10.0}
    s_net = net_scores.get(net_clean, 40.0)
    
    s_comp = min(complaints / 100.0, 1.0) * 100.0
    s_user = float(internet_users)
    s_fib = 100.0 if fiber_clean == "no" else 0.0
    
    # Weights
    w_sig = 0.20
    w_comp = 0.20
    w_pop = 0.15
    w_dist = 0.15
    w_net = 0.15
    w_fib = 0.10
    w_user = 0.05
    
    score = (w_pop * s_pop + 
             w_sig * s_sig + 
             w_dist * s_dist + 
             w_net * s_net + 
             w_comp * s_comp + 
             w_user * s_user + 
             w_fib * s_fib)
    
    score = round(score, 2)
    
    if score >= 75.0:
        cat = "High"
    elif score >= 45.0:
        cat = "Medium"
    else:
        cat = "Low"
        
    return score, cat

# PDF Helper Class for page numbering and headers
class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            super().showPage()
        super().save()

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 8)
        self.setFillColor(colors.HexColor("#666666"))
        
        # Header (on pages other than the first page)
        if self._pageNumber > 1:
            self.drawString(54, 750, "STIP - Smart Telecom Infrastructure Planner (BSNL Portal)")
            self.setStrokeColor(colors.HexColor("#D1D5DB"))
            self.setLineWidth(0.5)
            self.line(54, 742, 558, 742)
            
        # Footer
        footer_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(558, 40, footer_text)
        self.drawString(54, 40, "CONFIDENTIAL - BSNL INTERNAL PLANNING REPORT")
        self.setStrokeColor(colors.HexColor("#D1D5DB"))
        self.setLineWidth(0.5)
        self.line(54, 52, 558, 52)
        
        self.restoreState()

# Routes
@app.route('/')
@app.route('/index')
def index():
    return render_template('index.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password'].strip()
        
        conn = get_db_connection()
        user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        conn.close()
        
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            flash('Logged in successfully.', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Invalid username or password.', 'danger')
            
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out successfully.', 'info')
    return redirect(url_for('index'))

# Route wrapper to check login
def login_required(f):
    import functools
    @functools.wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/dashboard')
@login_required
def dashboard():
    conn = get_db_connection()
    # Stats
    total_areas = conn.execute("SELECT COUNT(*) FROM areas").fetchone()[0]
    high_count = conn.execute("SELECT COUNT(*) FROM areas WHERE category = 'High'").fetchone()[0]
    med_count = conn.execute("SELECT COUNT(*) FROM areas WHERE category = 'Medium'").fetchone()[0]
    low_count = conn.execute("SELECT COUNT(*) FROM areas WHERE category = 'Low'").fetchone()[0]
    avg_signal = conn.execute("SELECT AVG(signal_strength) FROM areas").fetchone()[0] or 0.0
    
    # Tables
    recent_areas = conn.execute("SELECT * FROM areas ORDER BY created_at DESC LIMIT 5").fetchall()
    top_priority = conn.execute("SELECT * FROM areas ORDER BY priority_score DESC LIMIT 5").fetchall()
    conn.close()
    
    return render_template('dashboard.html', 
                           total_areas=total_areas, 
                           high_count=high_count, 
                           med_count=med_count, 
                           low_count=low_count, 
                           avg_signal=round(avg_signal, 2),
                           recent_areas=recent_areas, 
                           top_priority=top_priority)

@app.route('/add', methods=['GET', 'POST'])
@login_required
def add_area():
    if request.method == 'POST':
        name = request.form['name'].strip()
        population = int(request.form['population'])
        signal_strength = int(request.form['signal_strength'])
        distance = float(request.form['distance'])
        internet_users = int(request.form['internet_users'])
        complaints = int(request.form['complaints'])
        fiber = request.form['fiber']
        network_type = request.form['network_type']
        
        # Calculate TIPI score
        score, cat = calculate_tipi(population, signal_strength, distance, internet_users, complaints, fiber, network_type)
        
        conn = get_db_connection()
        conn.execute('''
            INSERT INTO areas (name, population, signal_strength, distance, internet_users, complaints, fiber, network_type, priority_score, category)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (name, population, signal_strength, distance, internet_users, complaints, fiber, network_type, score, cat))
        conn.commit()
        conn.close()
        
        flash(f'Area "{name}" added successfully with Priority Score {score} ({cat}).', 'success')
        return redirect(url_for('view_areas'))
        
    return render_template('add_area.html', area=None)

@app.route('/edit/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_area(id):
    conn = get_db_connection()
    area = conn.execute("SELECT * FROM areas WHERE id = ?", (id,)).fetchone()
    
    if area is None:
        conn.close()
        flash('Area not found.', 'danger')
        return redirect(url_for('view_areas'))
        
    if request.method == 'POST':
        name = request.form['name'].strip()
        population = int(request.form['population'])
        signal_strength = int(request.form['signal_strength'])
        distance = float(request.form['distance'])
        internet_users = int(request.form['internet_users'])
        complaints = int(request.form['complaints'])
        fiber = request.form['fiber']
        network_type = request.form['network_type']
        
        # Recalculate
        score, cat = calculate_tipi(population, signal_strength, distance, internet_users, complaints, fiber, network_type)
        
        conn.execute('''
            UPDATE areas 
            SET name = ?, population = ?, signal_strength = ?, distance = ?, internet_users = ?, complaints = ?, fiber = ?, network_type = ?, priority_score = ?, category = ?
            WHERE id = ?
        ''', (name, population, signal_strength, distance, internet_users, complaints, fiber, network_type, score, cat, id))
        conn.commit()
        conn.close()
        
        flash(f'Area "{name}" updated successfully.', 'success')
        return redirect(url_for('view_areas'))
        
    conn.close()
    return render_template('add_area.html', area=area)

@app.route('/delete/<int:id>', methods=['POST'])
@login_required
def delete_area(id):
    conn = get_db_connection()
    conn.execute("DELETE FROM areas WHERE id = ?", (id,))
    conn.commit()
    conn.close()
    flash('Area record deleted.', 'success')
    return redirect(url_for('view_areas'))

@app.route('/view')
@login_required
def view_areas():
    search = request.args.get('search', '').strip()
    net_type = request.args.get('network_type', '').strip()
    category = request.args.get('category', '').strip()
    
    query = "SELECT * FROM areas WHERE 1=1"
    params = []
    
    if search:
        query += " AND name LIKE ?"
        params.append(f"%{search}%")
    if net_type:
        query += " AND network_type = ?"
        params.append(net_type)
    if category:
        query += " AND category = ?"
        params.append(category)
        
    query += " ORDER BY created_at DESC"
    
    conn = get_db_connection()
    areas = conn.execute(query, params).fetchall()
    conn.close()
    
    return render_template('view_area.html', areas=areas, search=search, network_type=net_type, category=category)

@app.route('/calculate', methods=['POST'])
@login_required
def calculate_priority():
    conn = get_db_connection()
    areas = conn.execute("SELECT * FROM areas").fetchall()
    
    for area in areas:
        score, cat = calculate_tipi(
            area['population'], 
            area['signal_strength'], 
            area['distance'], 
            area['internet_users'], 
            area['complaints'], 
            area['fiber'], 
            area['network_type']
        )
        conn.execute("UPDATE areas SET priority_score = ?, category = ? WHERE id = ?", (score, cat, area['id']))
        
    conn.commit()
    conn.close()
    flash('TIPI Scores re-calculated for all areas successfully.', 'success')
    return redirect(url_for('ranking'))

@app.route('/ranking')
@login_required
def ranking():
    conn = get_db_connection()
    areas = conn.execute("SELECT * FROM areas ORDER BY priority_score DESC").fetchall()
    conn.close()
    
    # Dynamic BSNL standard recommendations based on priority category and specific gaps
    ranked_areas = []
    for rank, area in enumerate(areas, 1):
        rec = ""
        if area['category'] == 'High':
            if area['fiber'].strip().lower() == 'no' and area['network_type'] in ['2G', '3G']:
                rec = "CRITICAL: Deploy fiber backhaul & replace legacy BTS with multi-sector 5G node immediately."
            elif area['complaints'] > 80:
                rec = "URGENT: Capacity expansion. Add micro-cells to distribute load and resolve high complaints."
            else:
                rec = "IMMEDIATE: Erect new cell tower and deploy 4G/5G co-location to resolve poor coverage."
        elif area['category'] == 'Medium':
            if area['distance'] > 8.0:
                rec = "SCHEDULED: Install macro-repeater or optimize tilt on current towers to bridge range."
            elif area['network_type'] == '3G':
                rec = "SCHEDULED: Upgrade transceiver modules to LTE (4G) during next cycle."
            else:
                rec = "ROUTINE: Increase RF power output, perform sector re-routing, and check user complaints."
        else:
            rec = "MONITOR: Core KPIs are satisfactory. Scheduled preventive maintenance check every 6 months."
            
        ranked_areas.append({
            'rank': rank,
            'id': area['id'],
            'name': area['name'],
            'population': area['population'],
            'network_type': area['network_type'],
            'fiber': area['fiber'],
            'priority_score': area['priority_score'],
            'category': area['category'],
            'recommendation': rec
        })
        
    return render_template('ranking.html', areas=ranked_areas)

@app.route('/analytics')
@login_required
def analytics():
    return render_template('analytics.html')

# Endpoint returning JSON data for Chart.js
@app.route('/api/chart-data')
@login_required
def chart_data():
    conn = get_db_connection()
    areas_db = conn.execute("SELECT * FROM areas ORDER BY priority_score DESC").fetchall()
    conn.close()
    
    # Convert Row objects to list of dicts
    areas = [dict(row) for row in areas_db]
    
    # 1. Top 10 Priority Areas
    top_10 = [{'name': a['name'], 'score': a['priority_score'], 'category': a['category']} for a in areas[:10]]
    
    # 2. Complaint distribution (group by network type or category or area)
    complaints = [{'name': a['name'], 'complaints': a['complaints']} for a in areas]
    
    # 3. Population vs Priority Scatter
    scatter = [{'x': a['population'], 'y': a['priority_score'], 'label': a['name']} for a in areas]
    
    # 4. Network type distribution
    net_types = {}
    for a in areas:
        net = a['network_type']
        net_types[net] = net_types.get(net, 0) + 1
        
    # 5. Category distribution
    cat_counts = {"High": 0, "Medium": 0, "Low": 0}
    for a in areas:
        cat = a['category']
        if cat in cat_counts:
            cat_counts[cat] += 1
            
    return jsonify({
        'top_10': top_10,
        'complaints': complaints,
        'scatter': scatter,
        'net_types': net_types,
        'categories': cat_counts
    })

# Excel Upload Handler
@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    if 'file' not in request.files:
        flash('No file part', 'danger')
        return redirect(url_for('dashboard'))
        
    file = request.files['file']
    if file.filename == '':
        flash('No selected file', 'danger')
        return redirect(url_for('dashboard'))
        
    if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            # Read Excel via Pandas
            df = pd.read_excel(filepath)
            
            # Column mapping (tolerant of case and spaces)
            col_map = {
                'area name': 'name', 'area': 'name', 'name': 'name',
                'population': 'population', 'pop': 'population',
                'signal strength': 'signal_strength', 'signal': 'signal_strength', 'signalstrength': 'signal_strength',
                'distance': 'distance', 'distance (km)': 'distance', 'distance from tower': 'distance',
                'internet users': 'internet_users', 'internet users %': 'internet_users', 'internetusers': 'internet_users',
                'complaints': 'complaints', 'complaint count': 'complaints', 'complaints count': 'complaints',
                'fiber': 'fiber', 'fiber available': 'fiber', 'fiber available (yes/no)': 'fiber',
                'network type': 'network_type', 'network': 'network_type', 'networktype': 'network_type'
            }
            
            # Clean dataframe columns
            cleaned_cols = {}
            for col in df.columns:
                col_lower = str(col).strip().lower()
                if col_lower in col_map:
                    cleaned_cols[col] = col_map[col_lower]
                    
            if len(set(cleaned_cols.values())) < 8:
                missing = set(['name', 'population', 'signal_strength', 'distance', 'internet_users', 'complaints', 'fiber', 'network_type']) - set(cleaned_cols.values())
                flash(f'Excel file lacks required columns. Missing: {list(missing)}. Please use the download template.', 'danger')
                return redirect(url_for('dashboard'))
                
            df = df.rename(columns=cleaned_cols)
            df = df[['name', 'population', 'signal_strength', 'distance', 'internet_users', 'complaints', 'fiber', 'network_type']]
            
            conn = get_db_connection()
            rows_inserted = 0
            
            for index, row in df.iterrows():
                name = str(row['name']).strip()
                if not name or name == 'nan':
                    continue
                
                # Coerce types
                try:
                    pop = int(row['population'])
                    sig = int(row['signal_strength'])
                    dist = float(row['distance'])
                    users = int(row['internet_users'])
                    comp = int(row['complaints'])
                    fib = str(row['fiber']).strip()
                    net = str(row['network_type']).strip()
                except Exception as e:
                    # Skip rows with format errors
                    continue
                    
                score, cat = calculate_tipi(pop, sig, dist, users, comp, fib, net)
                
                conn.execute('''
                    INSERT INTO areas (name, population, signal_strength, distance, internet_users, complaints, fiber, network_type, priority_score, category)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (name, pop, sig, dist, users, comp, fib, net, score, cat))
                rows_inserted += 1
                
            conn.commit()
            conn.close()
            
            # Cleanup file
            os.remove(filepath)
            flash(f'Successfully imported {rows_inserted} areas from Excel file. All priority indexes recalculated.', 'success')
            
        except Exception as e:
            flash(f'Error reading Excel file: {str(e)}', 'danger')
            if os.path.exists(filepath):
                os.remove(filepath)
                
        return redirect(url_for('dashboard'))
    else:
        flash('Invalid file extension. Please upload a standard Excel file (.xlsx or .xls).', 'danger')
        return redirect(url_for('dashboard'))

# Template Downloader for Excel Bulk Upload
@app.route('/download/template')
@login_required
def download_template():
    # Generate an empty template with columns
    wb = Workbook()
    ws = wb.active
    ws.title = "STIP Import Template"
    
    headers = [
        "Area Name", "Population", "Signal Strength", 
        "Distance", "Internet Users", "Complaints", 
        "Fiber", "Network Type"
    ]
    
    # Styles
    header_fill = PatternFill(start_color="0B2545", end_color="0B2545", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")
    
    ws.append(headers)
    
    # Sample Row
    ws.append([
        "Sample Circle A", 45000, 3, 
        5.4, 60, 45, 
        "No", "3G"
    ])
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 15)
        
    # Apply header format
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    temp_path = os.path.join(app.config['UPLOAD_FOLDER'], 'stip_upload_template.xlsx')
    wb.save(temp_path)
    
    return send_file(temp_path, as_attachment=True, download_name="stip_bulk_upload_template.xlsx")

@app.route('/reports')
@login_required
def reports():
    conn = get_db_connection()
    areas = conn.execute("SELECT * FROM areas ORDER BY priority_score DESC").fetchall()
    total = len(areas)
    high = conn.execute("SELECT COUNT(*) FROM areas WHERE category='High'").fetchone()[0]
    med = conn.execute("SELECT COUNT(*) FROM areas WHERE category='Medium'").fetchone()[0]
    low = conn.execute("SELECT COUNT(*) FROM areas WHERE category='Low'").fetchone()[0]
    conn.close()
    
    return render_template('report.html', areas=areas, total=total, high=high, med=med, low=low)

# Report Exports - EXCEL
@app.route('/report/excel')
@login_required
def export_excel():
    conn = get_db_connection()
    areas = conn.execute("SELECT * FROM areas ORDER BY priority_score DESC").fetchall()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Priority Analysis Report"
    
    # Headers
    headers = [
        "Rank", "Circle/Area Name", "Population", "Signal Strength (1-5)", 
        "Distance to Tower (km)", "Internet Users %", "Monthly Complaints", 
        "Fiber Backhaul", "Current Tech", "TIPI Priority Score", "Category"
    ]
    ws.append(headers)
    
    # Rows
    for rank, area in enumerate(areas, 1):
        ws.append([
            rank, area['name'], area['population'], area['signal_strength'],
            area['distance'], area['internet_users'], area['complaints'],
            area['fiber'], area['network_type'], area['priority_score'], area['category']
        ])
        
    # Style Sheet formatting
    header_fill = PatternFill(start_color="0B2545", end_color="0B2545", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    # Category fills
    high_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid") # soft red
    med_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # soft yellow
    low_fill = PatternFill(start_color="D1E7DD", end_color="D1E7DD", fill_type="solid")  # soft green
    
    border_thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    # Apply cell styles
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 3, 12)
        
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    for row_idx in range(2, len(areas) + 2):
        category = ws.cell(row=row_idx, column=11).value
        # Borders & Fonts
        for col_idx in range(1, 12):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border_thin
            if col_idx in [1, 3, 4, 5, 6, 7, 8, 9, 10]:
                cell.alignment = Alignment(horizontal="center")
            
            # Apply color highlights based on category
            if category == 'High':
                cell.fill = high_fill
            elif category == 'Medium':
                cell.fill = med_fill
            elif category == 'Low':
                cell.fill = low_fill
                
    # Freeze pane
    ws.freeze_panes = "A2"
    
    file_path = os.path.join(REPORTS_FOLDER, 'stip_infrastructure_priority_report.xlsx')
    wb.save(file_path)
    
    return send_file(file_path, as_attachment=True, download_name=f"STIP_Priority_Report_{datetime.now().strftime('%Y%m%d')}.xlsx")

# Report Exports - PDF using ReportLab
@app.route('/report/pdf')
@login_required
def export_pdf():
    conn = get_db_connection()
    areas = conn.execute("SELECT * FROM areas ORDER BY priority_score DESC").fetchall()
    
    total = len(areas)
    high = conn.execute("SELECT COUNT(*) FROM areas WHERE category='High'").fetchone()[0]
    med = conn.execute("SELECT COUNT(*) FROM areas WHERE category='Medium'").fetchone()[0]
    low = conn.execute("SELECT COUNT(*) FROM areas WHERE category='Low'").fetchone()[0]
    conn.close()
    
    pdf_path = os.path.join(REPORTS_FOLDER, 'stip_report.pdf')
    
    # 0.75 in margins = 54 pt
    doc = SimpleDocTemplate(pdf_path, pagesize=letter, leftMargin=54, rightMargin=54, topMargin=54, bottomMargin=54)
    story = []
    
    # Style Sheet
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=22,
        leading=26,
        textColor=colors.HexColor('#0B2545'),
        alignment=0, # Left-aligned
        spaceAfter=6
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubtitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=12,
        leading=16,
        textColor=colors.HexColor('#666666'),
        alignment=0,
        spaceAfter=20
    )
    
    h1_style = ParagraphStyle(
        'Header1',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=14,
        leading=18,
        textColor=colors.HexColor('#134074'),
        spaceBefore=14,
        spaceAfter=8,
        keepWithNext=True
    )
    
    body_style = ParagraphStyle(
        'BodyTextCustom',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9.5,
        leading=13.5,
        textColor=colors.HexColor('#333333'),
        spaceAfter=8
    )
    
    cell_style = ParagraphStyle(
        'TableCell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#333333'),
        alignment=1 # Centered
    )
    
    cell_left_style = ParagraphStyle(
        'TableCellLeft',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#333333'),
        alignment=0 # Left
    )
    
    header_cell_style = ParagraphStyle(
        'TableHeaderCell',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.white,
        alignment=1 # Centered
    )
    
    # Add Cover Title
    story.append(Paragraph("Smart Telecom Infrastructure Planner (STIP)", title_style))
    story.append(Paragraph(f"Decision Support Report - Generated on {datetime.now().strftime('%B %d, %Y at %H:%M')}", subtitle_style))
    story.append(Spacer(1, 10))
    
    # Exec Summary Card
    story.append(Paragraph("Executive Summary", h1_style))
    summary_text = (
        f"This strategic planning document evaluates {total} geographical circles to prioritize BSNL's "
        "infrastructure expansion and resource deployment. Scores are calculated via the Telecom Infrastructure "
        f"Priority Index (TIPI) algorithm. Currently, {high} circles are flagged as High Priority "
        f"(immediate installation recommended), {med} as Medium Priority (scheduled upgrades required), "
        f"and {low} as Low Priority (routine maintenance)."
    )
    story.append(Paragraph(summary_text, body_style))
    story.append(Spacer(1, 10))
    
    # Summary Table
    summary_data = [
        [Paragraph("Category", header_cell_style), Paragraph("Threshold", header_cell_style), Paragraph("Circles Identified", header_cell_style), Paragraph("Recommended Action", header_cell_style)],
        [Paragraph("High Priority", cell_left_style), Paragraph("TIPI >= 75", cell_style), Paragraph(str(high), cell_style), Paragraph("Immediate Fiber Backhaul / 5G Installation", cell_left_style)],
        [Paragraph("Medium Priority", cell_left_style), Paragraph("45 <= TIPI < 75", cell_style), Paragraph(str(med), cell_style), Paragraph("Upgrade legacy node to 4G LTE / Optimize Tilt", cell_left_style)],
        [Paragraph("Low Priority", cell_left_style), Paragraph("TIPI < 45", cell_style), Paragraph(str(low), cell_style), Paragraph("Preventive maintenance / Routine KPI monitoring", cell_left_style)],
    ]
    
    summary_table = Table(summary_data, colWidths=[100, 80, 100, 224])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0B2545')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('TOPPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D1D5DB')),
        ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#FDF2F2')), # soft red
        ('BACKGROUND', (0,2), (-1,2), colors.HexColor('#FEFCE8')), # soft yellow
        ('BACKGROUND', (0,3), (-1,3), colors.HexColor('#F0FDF4')), # soft green
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.white]),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Priority Rankings Table
    story.append(Paragraph("Infrastructure Priority Index Rankings", h1_style))
    
    table_data = [
        [
            Paragraph("Rk", header_cell_style), 
            Paragraph("Circle/Area Name", header_cell_style), 
            Paragraph("Pop.", header_cell_style), 
            Paragraph("Sig.", header_cell_style), 
            Paragraph("Dist.", header_cell_style), 
            Paragraph("Comp.", header_cell_style), 
            Paragraph("Fib.", header_cell_style), 
            Paragraph("Tech", header_cell_style), 
            Paragraph("TIPI", header_cell_style), 
            Paragraph("Priority", header_cell_style)
        ]
    ]
    
    for r, area in enumerate(areas, 1):
        # Color matching
        cat = area['category']
        if cat == 'High':
            cat_color = colors.HexColor('#991B1B') # dark red
        elif cat == 'Medium':
            cat_color = colors.HexColor('#9A3412') # dark orange
        else:
            cat_color = colors.HexColor('#166534') # dark green
            
        cat_p = Paragraph(f"<font color='{cat_color.hexval()}'><b>{cat}</b></font>", cell_style)
        
        table_data.append([
            Paragraph(str(r), cell_style),
            Paragraph(area['name'], cell_left_style),
            Paragraph(f"{area['population']:,}", cell_style),
            Paragraph(str(area['signal_strength']), cell_style),
            Paragraph(f"{area['distance']} km", cell_style),
            Paragraph(str(area['complaints']), cell_style),
            Paragraph(area['fiber'], cell_style),
            Paragraph(area['network_type'], cell_style),
            Paragraph(f"<b>{area['priority_score']}</b>", cell_style),
            cat_p
        ])
        
    rankings_table = Table(table_data, colWidths=[20, 140, 48, 28, 38, 38, 28, 32, 42, 50])
    
    # Base styling
    ts = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#134074')),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,0), 5),
        ('TOPPADDING', (0,0), (-1,0), 5),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
        ('PADDING', (0,0), (-1,-1), 4),
    ])
    
    # Add alternating colors to data rows
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            ts.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#F9FAFB'))
            
    rankings_table.setStyle(ts)
    story.append(rankings_table)
    
    # Build Document using NumberedCanvas for header/footer
    doc.build(story, canvasmaker=NumberedCanvas)
    
    return send_file(pdf_path, as_attachment=True, download_name=f"STIP_Priority_Report_{datetime.now().strftime('%Y%m%d')}.pdf")

# Database initializer check in Flask
@app.before_request
def check_db():
    if not os.path.exists(DB_PATH):
        # Trigger initialization
        from db_init import init_db
        init_db()

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
